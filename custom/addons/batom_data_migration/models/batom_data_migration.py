# -*- coding: utf-8 -*-
# Copyright <2016> <Batom Co., Ltd.>
# License AGPL-3.0 or later (http://www.gnu.org/licenses/agpl).

import os
import logging
import re
from datetime import datetime, timedelta
import time
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import coordinate_from_string, column_index_from_string
from openpyxl.utils.cell import rows_from_range
from odoo import models, fields, api,  _
import odoo
import odoo.addons.base_external_dbsource

_logger = logging.getLogger(__name__)
    
_currencyMapping = ({
    'NTD': 'TWD',
    'GRP': 'GBP',
    })
def _currencyIdConversion(self, currencyId):
    returnedCurrencyId = None
    if currencyId and currencyId.strip():
        if currencyId in _currencyMapping:
            currencyId = _currencyMapping[currencyId]
        currencyIds = self.env['res.currency'].search([('name', '=', currencyId)]).ids
        if len(currencyIds) > 0:
            returnedCurrencyId = currencyIds.pop(0)
    return returnedCurrencyId

#BX
#P C
#SET
#桶
#EA
#ＰＣ
#英呎
#M
#PCS
#Kg

_uomMapping = ({
    'Kg': 'kg',
    'M': 'm',
    'P C': 'PC',
    'pc': 'PC',
    'pcs': 'PCS',
    u'ＰＣ': 'PC',
    'set': 'SET',
    })
def _uomIdConversion(self, uom):
    uomModel = self.env['product.uom']
    uomCategoryModel = self.env['product.uom.categ']
    uomId = 1 # uom cannot be null
    if uom and uom.strip():
        uom = uom.strip()
        if uom in _uomMapping:
            uom = _uomMapping[uom]
        uomIds = uomModel.search([('name', '=', uom)]).ids
        if len(uomIds) > 0:
            uomId = uomIds.pop(0)
        else:
            factor = 1
            rounding = 0.01
            uom_type = 'bigger'
            if uom == u'桶':
                category_id = uomCategoryModel.with_context(lang='en_US').search([('name', '=', 'Volume')])[0].id
            else:
                category_id = uomCategoryModel.with_context(lang='en_US').search([('name', '=', 'Unit')])[0].id
            uomValues = ({
                'name': uom,
                'category_id': category_id,
                'factor': factor,
                'rounding': rounding,
                'uom_type': uom_type,
                })
            odooUom = uomModel.create(uomValues)
            if uom == u'桶':
                _updateTranslation(self, 'product.uom,name', odooUom.id, 'barrel', uom)
            
    return uomId
    
def _updateTranslation(self, name, res_id, src, value):
    lang = 'zh_TW'
    type = 'model'
    state = 'translated'
    translationModel = self.env['ir.translation']
    if not value or not value.strip():
        value = src
    values = ({
        'lang': lang,
        'type': type,
        'name': name,
        'res_id': res_id,
        'src': src,
        'value': value,
        'state': state,
        })
    translations = translationModel.search([
        ('lang', '=', lang),
        ('type', '=', type),
        ('name', '=', name),
        ('res_id', '=', res_id),
        ('src', '=', value),
        ])
    if len(translations) == 0:
        translationModel.create(values)
    else:
        translations[0].write(values)
        
def _getSupplier(self, supplierCode, defaultIfNotFound):
    supplier = None
    if supplierCode != None:
        suppliers = self.env['res.partner'].search([
            ('supplier', '=', True),
            ('x_supplier_code', '=', supplierCode)
            ])
    else:
        suppliers = []
    
    if len(suppliers) > 0:
        supplier =  suppliers[0]
    elif defaultIfNotFound:
        supplierCompany = self.env['res.company']._company_default_get()
        if supplierCompany:
            supplier = supplierCompany.partner_id
            if not supplier.x_supplier_code:
                supplier.write({
                    'x_supplier_code': '01',
                    })

    return supplier
        
def _getProduct(self, productCode):
    product = None
    products = self.env['product.product'].search([
        ('default_code', '=', productCode)
        ])
    if len(products) > 0:
        product =  products[0]

    return product
        
def _getProductTemplate(self, productCode):
    product = None
    products = self.env['product.template'].search([
        ('default_code', '=', productCode)
        ])
    if len(products) > 0:
        product =  products[0]

    return product

def _getWorkcenter(self, processCode, supplierCode, createIfNotExist):
    workcenter = None
    try:
        process = self.env['product.template'].search([
            ('x_is_process', '=', True),
            ('default_code', '=', processCode)
            ])[0]
        supplier = _getSupplier(self, supplierCode, True)
        workcenterModel = self.env['mrp.workcenter']
        workcenters = workcenterModel.search([
            ('x_supplier_id', '=', supplier.id),
            ('x_process_id', '=', process.id)
            ])
        if len(workcenters) > 0:
            workcenter =  workcenters[0]
        elif createIfNotExist:
            workcenterValues = ({
                'name': process.name + ' <- ' + supplier.display_name,
                'code': processCode + ' <- ' + (supplier.x_supplier_code if supplier.x_supplier_code else '01'),
                'x_process_id': process.id,
                'x_supplier_id': supplier.id,
                'resource_type': 'material',
                })
            workcenter = workcenterModel.create(workcenterValues);
    except Exception:
        _logger.warning('Exception in migrate_bom:', exc_info=True)
        import pdb; pdb.set_trace()

    return workcenter

def _createOdooProduct(self, cursorChi, productId):
    odooProduct = False
    chiProduct = cursorChi.execute(u"SELECT ProdId, ClassId, ProdForm, Unit, ProdName, EngName, ProdDesc, "
        u"CurrId, CAvgCost, SuggestPrice, NWeight, NUnit "
        u"FROM comProduct "
        u"WHERE ProdId = '" + productId.decode('utf-8') + u"'"
        ).fetchone()
    if chiProduct != None:
        if chiProduct.ProdName and chiProduct.ProdName.strip():
            name = chiProduct.ProdName
        else:
            name = chiProduct.ProdId
        currency_id = _currencyIdConversion(self, chiProduct.CurrId)
        uom_id = _uomIdConversion(self, chiProduct.Unit.decode('utf-8'))
        # ProdForm: 1-物料，2半成品，3-成品，4-採購件，5-組合品，6-非庫存品，7-非庫存品(管成本)，8-易耗品
        # ClassId ClassName
        # --------------
        # *	特殊科目
        # 1	運費
        # 2	雜項支出
        # 3	包裝費
        # 4	樣品費
        # 5	製-包裝費
        # 6	進料
        # 7	製-模具費
        # A	原料
        # B	半成品
        # C	成品
        # D	零配件
        # E	物料
        # F	模治具
        # G	紙箱
        # H	開發件
        # I	商品
        sale_ok = False
        purchase_ok = False
        type = 'consu' # 'consu', 'service', 'product'
        tracking = 'none'
        if chiProduct.ProdForm == 3 or chiProduct.ClassId == 'C' or chiProduct.ClassId == 'I':
            sale_ok = True
        if chiProduct.ProdForm in [1, 2, 4]:
            purchase_ok = True
        if chiProduct.ProdForm <= 5:
            type = 'product'
            tracking = 'lot'
        productValues = ({
            'name': name,
            'default_code': chiProduct.ProdId,
            'x_saved_code': chiProduct.ProdId,
            'type': type,
            'tracking': tracking,
            'description': chiProduct.ProdDesc,
            'sale_ok': sale_ok,
            'purchase_ok': purchase_ok,
            'currency_id': currency_id,
            'standard_price': chiProduct.CAvgCost,
            'price': chiProduct.SuggestPrice,
            'uom_id': uom_id,
            'uom_po_id': uom_id,
            'weight': chiProduct.NWeight,
            # warehouse_id, location_id, routes_id,
            })
        
        productModel = self.env['product.product']
        odooProducts = productModel.search([('default_code', '=', chiProduct.ProdId)])
        if len(odooProducts) == 0:
            odooProduct = productModel.create(productValues)
        else:
            odooProduct = odooProducts[0]
            odooProduct.write(productValues)
        if chiProduct.EngName and chiProduct.EngName.strip():
            engName = chiProduct.EngName
            _updateTranslation(self, 'product.template,name', odooProduct.product_tmpl_id.id, engName, name)
    return odooProduct

def _createOdooBom(self, cursorChi, chiBom, itemNo, active):
    bomModel = self.env['mrp.bom']
    productTemplateModel = self.env['product.template']
    template = productTemplateModel.search([('default_code', '=', chiBom.ProductId)])[0]
    chiBomMaterials = cursorChi.execute(
        u"SELECT SerNo, SubProdId, QtyOfBatch "
        u"FROM prdBOMMats "
        u"WHERE ProductId='" + chiBom.ProductId.decode('utf-8') + u"' and ItemNo=" + str(chiBom.ItemNo) + u" "
        u"ORDER BY SerNo").fetchall()
    chiBomProcesses = cursorChi.execute(
        u"SELECT SerNo, MkPgmId, ProdtClass, Producer, DailyProdtQty, PrepareDays, WorkTimeOfBatch, PriceOfProc "
        u"FROM prdBOMPgms "
        u"WHERE MainProdId='" + chiBom.ProductId.decode('utf-8') + u"' and ItemNo=" + str(chiBom.ItemNo) + u" "
        u"ORDER BY SerNo").fetchall()
    odooRouting = self._createRouting(chiBom, chiBomMaterials, chiBomProcesses, itemNo, active)
    bomValues = ({
        'code': u'~' + chiBom.ProductId.decode('utf-8') + u"#" + str(itemNo),
        'product_tmpl_id': template.id,
        'x_batom_bom_no': itemNo,
        'x_version_description': chiBom.CurVersion,
        'product_qty': chiBom.BatchAmount,
        'type': 'normal',
        'routing_id': odooRouting.id if odooRouting else False,
        'active': active,
        })
    
    odooBoms = bomModel.search([
        ('product_tmpl_id', '=', template.id),
        ('x_batom_bom_no', '=', itemNo),
        ])
    if len(odooBoms) == 0:
        odooBom = bomModel.create(bomValues)
    else:
        odooBom = odooBoms[0]
        odooBom.write(bomValues)
    self._createBomLines(odooBom, odooRouting, chiBom, chiBomMaterials, chiBomProcesses)

class _partner_migration:
    def __init__(self, id, shortName, fullName):
        self.id = id
        self.shortName = shortName
        self.fullName = fullName

class _partner_address:
    def __init__(self, zip, address, contactName, title, phone, mobile, fax, email, memo):
        self.zip = zip
        self.city = None
        self.state_id = None
        self.street = address
        self.street2 = None
        self.contactName = contactName
        self.title = title
        self.phone = phone
        self.mobile = mobile
        self.fax = fax
        self.email = email
        self.memo = memo

class _material_assignment:
    def __init__(self, product_code, product_qty, production_code, inventory_flag, inventory_code, adjustment_code, material_cost, cut_cost, forge_cost, material_lot, dispatch_date, used=False):
        self.product_code = product_code
        self.product_qty = product_qty
        self.production_code = production_code
        self.inventory_flag = inventory_flag
        self.inventory_code = inventory_code
        self.adjustment_code = adjustment_code
        self.material_cost = material_cost
        self.cut_cost = cut_cost
        self.forge_cost = forge_cost
        self.material_lot = material_lot
        self.dispatch_date = dispatch_date
        self.used = used

class _tool_sheet_format:
    def __init__(self, special_format_name, column_names, original_column_names):
        self.special_format_name = special_format_name
        self.column_names = column_names
        self.original_column_names = original_column_names

class BatomPartnerMigrationRefresh(models.TransientModel):
    _name = "batom.partner_migration_refresh"
    _description = "Refresh Partner Data Migration Table"

    def _countryIdConversion(self, areaId):
        areaIds = ["01", "02", "03", "04", "05", "06"]
        countryIds = [229, 235, 233, 49, 105, 199]
        countryId = None
        try:
            idx = areaIds.index(areaId)
            countryId = countryIds[idx]
        except ValueError:
            countryId = None
            
        return countryId
    
    def _partnerCategoryConversion(self, type, classId):
        categoryIds = self.env['res.partner.category'].search([('x_flag', '=', type), ('x_category_code', '=', classId)]).ids
        categoryId = None
        if len(categoryIds) > 0:
            categoryId = categoryIds.pop(0)
        return categoryId
        
    def _contactDisplayName(self, type, companyName, contactName):
        try:
            if contactName != None and contactName and contactName.strip():
                name = companyName + u', ' + contactName.decode('utf-8')
            else:
                types = ['contact', 'invoice', 'delivery', 'other']
                names = [u'聯絡人', u'發票地址', u'送貨地址', u'其他地址']
                idx = types.index(type)
                name = companyName + u', ' + names[idx]
        except Exception:
            _logger.warning('Exception in refresh_partner_data:', exc_info=True)
            name = companyName
        return name
    
    def _createPartnerContact(self, type, parentPartner, address):
        partnerModel = self.env['res.partner']
        if address.mobile and address.mobile.strip():
            mobile = address.mobile
        else:
            mobile = parentPartner.mobile if (type == 'contact') else None
        if address.email and address.email.strip():
            email = address.email
        else:
            email = parentPartner.email
        newPartner = partnerModel.create({
            'parent_id': parentPartner.id,
            'type': type,
            'email': email,
            'fax': address.fax,
            'name': address.contactName,
            'commercial_company_name': parentPartner.commercial_company_name,
            'mobile': mobile,
            'phone': address.phone,
            'is_company': 0, # 1 if (type != 'contact') else 0,
            'customer': parentPartner.customer,
            'supplier': parentPartner.supplier,
            'zip': address.zip,
            'city': address.city,
            'state_id': address.state_id,
            'street': address.street,
            'street2': address.street2,
            'comment': address.memo,
            })
        newPartner.write({'display_name': self._contactDisplayName(type, parentPartner.display_name, address.contactName)})
     
    @api.multi
    def refresh_partner_data(self):
        try:
            self.ensure_one()
            dbBatom = self.env['base.external.dbsource'].search([('name', '=', 'Batom')])
            dbChi = self.env['base.external.dbsource'].search([('name', '=', 'CHIComp01')])
            batomCustomers = dbBatom.execute('SELECT Id, ShortName, FullName FROM Customer ORDER BY Id')
            chiCustomers = dbChi.execute('SELECT Id, ShortName, FullName FROM comCustomer WHERE Flag=1 ORDER BY Id')
            odooCustomerIds = self.env['res.partner'].search([('is_company', '=', True), ('customer', '=', True)], order='x_customer_code').ids
            odooPartnerMigration = self.env['batom.partner_migration']
            odooPartnerMigration.search([]).unlink() # delete all migration records
            batomCustomer = None
            chiCustomer = None
            odooCustomer = None
            while (
                    batomCustomer != None or chiCustomer != None or odooCustomer != None or
                    len(batomCustomers) > 0 or len(chiCustomers) > 0 or len(odooCustomerIds) > 0
                    ):
                if batomCustomer == None and len(batomCustomers) > 0:
                    customer = batomCustomers.pop(0)
                    batomCustomer = _partner_migration(customer.Id, customer.ShortName, customer.FullName)
                if chiCustomer == None and len(chiCustomers) > 0:
                    customer = chiCustomers.pop(0)
                    chiCustomer = _partner_migration(customer.Id, customer.ShortName, customer.FullName)
                if odooCustomer == None and len(odooCustomerIds) > 0:
                    customer = self.env['res.partner'].browse(odooCustomerIds.pop(0))
                    odooCustomer = _partner_migration(customer.x_customer_code, customer.display_name, customer.name)
                    
                minId = None
                if batomCustomer != None:
                    minId = batomCustomer.id
                if chiCustomer != None and (minId == None or chiCustomer.id < minId):
                    minId = chiCustomer.id
                if odooCustomer != None and (minId == None or odooCustomer.id < minId):
                    minId = odooCustomer.id

                if batomCustomer != None and batomCustomer.id == minId:
                    in_id = batomCustomer.id
                    in_short_name = batomCustomer.shortName
                    in_full_name = batomCustomer.fullName
                    batomCustomer = None
                else:
                    in_id = None
                    in_short_name = None
                    in_full_name = None
                if chiCustomer != None and chiCustomer.id == minId:
                    chi_id = chiCustomer.id
                    chi_short_name = chiCustomer.shortName
                    chi_full_name = chiCustomer.fullName
                    chiCustomer = None
                else:
                    chi_id = None
                    chi_short_name = None
                    chi_full_name = None
                if odooCustomer != None and odooCustomer.id == minId:
                    odoo_id = odooCustomer.id
                    odoo_short_name = odooCustomer.shortName
                    odoo_full_name = odooCustomer.fullName
                    odooCustomer = None
                else:
                    odoo_id = None
                    odoo_short_name = None
                    odoo_full_name = None

                odooPartnerMigration.create({
                    'type': 1,
                    'in_id': in_id,
                    'in_short_name': in_short_name,
                    'in_full_name': in_full_name,
                    'chi_id': chi_id, 
                    'chi_short_name': chi_short_name,
                    'chi_full_name': chi_full_name,
                    'odoo_id': odoo_id,
                    'odoo_short_name': odoo_short_name,    
                    'odoo_full_name': odoo_full_name,
                    })

            batomSuppliers = dbBatom.execute('SELECT Id, ShortName, FullName FROM Supplier ORDER BY Id')
            chiSuppliers = dbChi.execute('SELECT Id, ShortName, FullName FROM comCustomer where Flag=2 ORDER BY Id')
            odooSupplierIds = self.env['res.partner'].search([('is_company', '=', True), ('supplier', '=', True)], order='x_supplier_code').ids
            batomSupplier = None
            chiSupplier = None
            odooSupplier = None
            while (
                    batomSupplier != None or chiSupplier != None or odooSupplier != None or
                    len(batomSuppliers) > 0 or len(chiSuppliers) > 0 or len(odooSupplierIds) > 0
                    ):
                if batomSupplier == None and len(batomSuppliers) > 0:
                    supplier = batomSuppliers.pop(0)
                    batomSupplier = _partner_migration(supplier.Id, supplier.ShortName, supplier.FullName)
                if chiSupplier == None and len(chiSuppliers) > 0:
                    supplier = chiSuppliers.pop(0)
                    chiSupplier = _partner_migration(supplier.Id, supplier.ShortName, supplier.FullName)
                if odooSupplier == None and len(odooSupplierIds) > 0:
                    supplier = self.env['res.partner'].browse(odooSupplierIds.pop(0))
                    odooSupplier = _partner_migration(supplier.x_supplier_code, supplier.display_name, supplier.name)
                    
                minId = None
                if batomSupplier != None:
                    minId = batomSupplier.id
                if chiSupplier != None and (minId == None or chiSupplier.id < minId):
                    minId = chiSupplier.id
                if odooSupplier != None and (minId == None or odooSupplier.id < minId):
                    minId = odooSupplier.id
            
                if batomSupplier != None and batomSupplier.id == minId:
                    in_id = batomSupplier.id
                    in_short_name = batomSupplier.shortName
                    in_full_name = batomSupplier.fullName
                    batomSupplier = None
                else:
                    in_id = None
                    in_short_name = None
                    in_full_name = None
                if chiSupplier != None and chiSupplier.id == minId:
                    chi_id = chiSupplier.id
                    chi_short_name = chiSupplier.shortName
                    chi_full_name = chiSupplier.fullName
                    chiSupplier = None
                else:
                    chi_id = None
                    chi_short_name = None
                    chi_full_name = None
                if odooSupplier != None and odooSupplier.id == minId:
                    odoo_id = odooSupplier.id
                    odoo_short_name = odooSupplier.shortName
                    odoo_full_name = odooSupplier.fullName
                    odooSupplier = None
                else:
                    odoo_id = None
                    odoo_short_name = None
                    odoo_full_name = None
            
                odooPartnerMigration.create({
                    'type': 2,
                    'in_id': in_id,
                    'in_short_name': in_short_name,
                    'in_full_name': in_full_name,
                    'chi_id': chi_id, 
                    'chi_short_name': chi_short_name,
                    'chi_full_name': chi_full_name,
                    'odoo_id': odoo_id,
                    'odoo_short_name': odoo_short_name,    
                    'odoo_full_name': odoo_full_name,
                    })
            self.env.cr.commit()            
        except Exception:
            _logger.warning('Exception in refresh_partner_data:', exc_info=True)
            
        return {
            'view_type': 'form',
            'view_mode': 'tree,form',
            'res_model': 'batom.partner_migration',
            'target': 'current',
            'res_id': 'view_partner_migration_tree',
            'type': 'ir.actions.act_window'
        }
    
    @api.multi
    def apply_partner_data(self):
        try:
            self.ensure_one()
            dbBatom = self.env['base.external.dbsource'].search([('name', '=', 'Batom')])
            dbChi = self.env['base.external.dbsource'].search([('name', '=', 'CHIComp01')])
            odooPartnerMigration = self.env['batom.partner_migration'].search([('odoo_id', '=', None)])
            partnerModel = self.env['res.partner']
            
            nCount = len(odooPartnerMigration)
            nDone = 0
            for migration in odooPartnerMigration:
                try:
                    if migration.type == 1:
                        codeColumnName = 'x_customer_code'
                        customerValue = 1
                        supplierValue = 0
                    else:
                        codeColumnName = 'x_supplier_code'
                        customerValue = 0
                        supplierValue = 1
                    if migration.chi_id != None and migration.chi_id and migration.chi_id.strip():
                        sql = (
                            "SELECT Id, ShortName, FullName, LinkMan, LinkManProf, "
                            "AreaId, ClassId, CurrencyId, Email, FaxNo, InvoiceHead, TaxNo, "
                            "MobileTel, Telephone1, Telephone2, Telephone3, WebAddress "
                            "FROM comCustomer "
                            "WHERE Flag=" + str(migration.type) + " AND Id='" + migration.chi_id + "'"
                            )
                        chiPartner = dbChi.execute(sql).pop(0)
                        
                        if chiPartner != None:
                            sql = ("SELECT AddrOfInvo, "
                                "A1.ZipCode AS ZipCode1, A1.Address AS Address1, A1.LinkMan AS LinkMan1, A1.LinkManProf AS LinkManProf1, A1.Telephone AS Telephone1, A1.FaxNo AS FaxNo1, A1.Memo AS Memo1, "
                                "A2.ZipCode AS ZipCode2, A2.Address AS Address2, A2.LinkMan AS LinkMan2, A2.LinkManProf AS LinkManProf2, A2.Telephone AS Telephone2, A2.FaxNo AS FaxNo2, A2.Memo AS Memo2, "
                                "A3.ZipCode AS ZipCode3, A3.Address AS Address3, A3.LinkMan AS LinkMan3, A3.LinkManProf AS LinkManProf3, A3.Telephone AS Telephone3, A3.FaxNo AS FaxNo3, A3.Memo AS Memo3 "
                                "FROM comCustDesc AS C "
                                "LEFT JOIN comCustAddress A1 ON (A1.Flag = C.Flag AND A1.Id = C.Id AND A1.AddrId = C.DeliverAddrId) "
                                "LEFT JOIN comCustAddress A2 ON (A2.Flag = C.Flag AND A2.Id = C.Id AND A2.AddrId = C.AddrId) "
                                "LEFT JOIN comCustAddress A3 ON (A3.Flag = C.Flag AND A3.Id = C.Id AND A3.AddrId = C.EngAddrId) "
                                "WHERE C.Flag=" + str(migration.type) + " AND C.Id='" + migration.chi_id + "'"
                                )
                            addresses = dbChi.execute(sql).pop(0)
                            address = None
                            addressShipping = None
                            addressInvoice = None
                            addressOther = None
                            if addresses != None:
                                if addresses.Address1 != None:
                                    addressShipping = _partner_address(addresses.ZipCode1, addresses.Address1, addresses.LinkMan1, addresses.LinkManProf1, addresses.Telephone1, None, addresses.FaxNo1, None, addresses.Memo1)
                                    address = addressShipping
                                if addresses.Address2 != None:
                                    addressInvoice = _partner_address(addresses.ZipCode2, addresses.Address2, addresses.LinkMan2, addresses.LinkManProf2, addresses.Telephone2, None, addresses.FaxNo2, None, addresses.Memo2)
                                    if address == None:
                                        address = addressInvoice
                                if addresses.Address3 != None:
                                    addressOther = _partner_address(addresses.ZipCode3, addresses.Address3, addresses.LinkMan3, addresses.LinkManProf3, addresses.Telephone3, None, addresses.FaxNo3, None, addresses.Memo3)
                                    if address == None:
                                        address = addressOther
                            
                            categoryId = self._partnerCategoryConversion(migration.type, chiPartner.ClassId)
                            newPartner = partnerModel.create({
                                codeColumnName: chiPartner.Id,
                                'country_id': self._countryIdConversion(chiPartner.AreaId),
                                'category_id': [(4, categoryId)] if (categoryId != None) else None,
                                'property_purchase_currency_id': _currencyIdConversion(self, chiPartner.CurrencyId),
                                'email': chiPartner.Email,
                                'fax': chiPartner.FaxNo if (chiPartner.FaxNo != None and chiPartner.FaxNo and chiPartner.FaxNo.strip()) else (
                                    address.fax if (address != None) else None),
                                'name': chiPartner.ShortName,
                                'commercial_company_name': chiPartner.InvoiceHead if chiPartner.InvoiceHead else chiPartner.FullName,
                                'mobile': chiPartner.MobileTel,
                                'vat': chiPartner.TaxNo,
                                'phone': chiPartner.Telephone1 if (chiPartner.Telephone1 != None and chiPartner.Telephone1 and chiPartner.Telephone1.strip()) else (
                                    address.phone if address != None else None),
                                'x_phone2': chiPartner.Telephone2,
                                'x_phone3': chiPartner.Telephone3,
                                'website': None if not chiPartner.WebAddress else chiPartner.WebAddress,
                                'is_company': 1,
                                'customer': customerValue,
                                'supplier': supplierValue,
                                'zip': address.zip if address != None else None,
                                'city': address.city if address != None else None,
                                'state_id': address.state_id if address != None else None,
                                'street': address.street if address != None else None,
                                'street2': address.street2 if address != None else None,
                                })
                            #newPartner.write({'display_name': chiPartner.ShortName})
                            
                            if chiPartner.LinkMan != None and chiPartner.LinkMan and chiPartner.LinkMan.strip():
                                contact = _partner_address(None, None, chiPartner.LinkMan, chiPartner.LinkManProf, newPartner.phone, None, newPartner.fax, None, None)
                                self._createPartnerContact('contact', newPartner, contact)
                            if addressShipping != None:
                                self._createPartnerContact('delivery', newPartner, addressShipping)
                            if addressInvoice != None:
                                self._createPartnerContact('invoice', newPartner, addressInvoice)
                            if addressOther != None:
                               self._createPartnerContact('other', newPartner, addressOther)
                            linkMans = dbChi.execute(
                                "SELECT PersonName, ProfTitle, Telephone, Mobile, Email, FaxNo, Memo "
                                "FROM comLinkMan "
                                "WHERE Flag=" + str(migration.type) + " AND CustomId='" + chiPartner.Id + "'"
                                )
                            if len(linkMans) > 0:
                                for linkMan in linkMans:
                                    contact = _partner_address(None, None, linkMan.PersonName, linkMan.ProfTitle, linkMan.Telephone, linkMan.Mobile, linkMan.FaxNo, linkMan.Email, linkMan.Memo)
                                    self._createPartnerContact('contact', newPartner, contact)
                    elif migration.in_id != None and migration.in_id and migration.in_id.strip():
                        if migration.type == 1:
                            inPartner = dbBatom.execute(
                                "select Id, ShortName, FullName "
                                "from Customer where Id='" + migration.in_id + "'"
                                ).pop(0)
                        else:
                            inPartner = dbBatom.execute(
                                "select Id, ShortName, FullName "
                                "from Supplier where Id='" + migration.in_id + "'"
                                ).pop(0)
                        if inPartner != None:
                            newPartner = partnerModel.create({
                                codeColumnName: inPartner.Id,
                                'name': inPartner.FullName,
                                'is_company': 1,
                                'customer': customerValue,
                                'supplier': supplierValue,
                                })
                            newPartner.write({'display_name': inPartner.ShortName})
                    nDone += 1
                    if nDone % 10 == 0:
                        print str(nDone) + '/' + str(nCount)
                        self.env.cr.commit()
                except Exception:
                    _logger.warning('Exception in apply_partner_data:', exc_info=True)
                    continue
            self.env.cr.commit()            
        except Exception:
            _logger.warning('Exception in apply_partner_data:', exc_info=True)
            
        self.refresh_partner_data()
        return {
            'view_type': 'form',
            'view_mode': 'tree,form',
            'res_model': 'batom.partner_migration',
            'target': 'current',
            'res_id': 'view_partner_migration_tree',
            'type': 'ir.actions.act_window'
        }
    
class BatomCreatSupplierWarehouses(models.TransientModel):
    _name = "batom.create_supplier_warehouses"
    _description = "Create Supplier Warehouses"
    root_location_id = fields.Integer(string='Root Location Id', required=True)

    @api.multi
    def create_supplier_warehouses(self):
        this = self[0]
        location_id = this.root_location_id
        try:
            self.ensure_one()
            locationModel = self.env['stock.location']
            supplierRootLocation = locationModel.browse(location_id)
            if supplierRootLocation != None:
                partnerLocations = locationModel.search([
                    '|', ('active', '=', True), ('active', '=', False),
                    ('partner_id', '!=', None)
                    ], ['partner_id'])
                supplierIdsWithLocations = []
                for partnerLocation in partnerLocations:
                    supplierIdsWithLocations.append(partnerLocation.partner_id)
                suppliers = self.env['res.partner'].search([
                    ('supplier', '=', True),
                    ('x_supplier_code', '!=', None),
                    ('id', 'not in', supplierIdsWithLocations),
                    ])
                for supplier in suppliers:
                    try:
                        name = supplier.display_name + ' (' + supplier.x_supplier_code + ')'
                        newLocation = locationModel.create({
                            'name': name,
                            'partner_id': supplier.id,
                            'location_id': supplierRootLocation.id,
                            'usage': 'supplier',
                        })
                    except Exception:
                        _logger.warning('Exception in create_supplier_warehouses:', exc_info=True)
                self.env.cr.commit()            
        except Exception:
            _logger.warning('Exception in create_supplier_warehouses:', exc_info=True)
    
class BatomMigrateChartOfAccount(models.TransientModel):
    _name = "batom.migrate_chart_of_account"
    _description = "Migrate Chart of Account"

    def _accountTypeIdConversion(self, chiSubClsId, chiSubjectId):
        returnedUserTypeId = None
        if chiSubClsId and chiSubClsId.strip():
            if chiSubClsId == '11':
                if int(chiSubjectId) < 1130000:
                    returnedUserTypeId = 3
                elif int(chiSubjectId) <= 1179000:
                    returnedUserTypeId = 1
            else:
                userTypeIdMap = ({
                    "12": 5,
                    "14": 6,
                    "15": 8,
                    "16": 6,
                    "17": 6,
                    "18": 6,
                    "21": 9,
                    "22": 9,
                    "25": 10,
                    "28": 10,
                    "31": 11,
                    "32": 11,
                    "33": 11,
                    "41": 14,
                    "42": 14,
                    "43": 14,
                    "46": 14,
                    "51": 17,
                    "52": 17,
                    "53": 17,
                    "54": 17,
                    "55": 17,
                    "56": 17,
                    "57": 17,
                    "59": 17,
                    "61": 16,
                    "62": 16,
                    "63": 16,
                    "71": 13,
                    "73": 16,
                    "81": 16,
                    "91": 17,
                    "92": 17,
                    })
                try:
                    returnedUserTypeId = userTypeIdMap[chiSubClsId]
                except Exception:
                    _logger.warning('Exception in migrate_chart_of_account:', exc_info=True)
                
        return returnedUserTypeId
        
    def _odooAccountIdFromCode(self, code):
        accountId = None
        if code and code.strip():
            accountIds = self.env['account.account'].search([('code', '=', code)]).ids
            if len(accountIds) > 0:
                accountId = accountIds.pop(0)
        return accountId
        
    @api.multi
    def migrate_chart_of_account(self):
        try:
            self.ensure_one()
            dbChi = self.env['base.external.dbsource'].search([('name', '=', 'CHIComp01')])
            batomAccountTypes = self.env['batom.account.type'].search([])
            if len(batomAccountTypes) == 0:
                chiAccountTypes = dbChi.execute('SELECT SubClsId, SubClsName FROM comSubjectCls ORDER BY SubClsId')
                while len(chiAccountTypes) > 0:
                    try:
                        chiAccountType = chiAccountTypes.pop(0)
                        self.env['batom.account.type'].create({
                            'code': chiAccountType.SubClsId,
                            'name': chiAccountType.SubClsName
                            })
                    except Exception:
                        _logger.warning('Exception in migrate_chart_of_account:', exc_info=True)
                        continue
                batomAccountTypes = self.env['batom.account.type'].search([])
                self.env.cr.commit()            
            
            batomAccountTypeIds = {}
            for batomAccountType in batomAccountTypes:
                batomAccountTypeIds[batomAccountType.code] = batomAccountType.id
            # base_external_dbsource seems to have size limit of the query result
            # query the values directly may not returns all qualified records
            # for some reason, '1441000' cannot be retrieved.  ignore it for now
            chiAccountIds = dbChi.execute('SELECT SubjectId FROM comSubject ORDER BY SubjectId')
            accountModel = self.env['account.account']
            for chiAccountId in chiAccountIds:
                try:
                    chiAccounts = dbChi.execute("SELECT SubClsId, SubjectId, SubjectName, ParentSubId, CurrId, Description FROM comSubject WHERE SubjectId = '" + chiAccountId.SubjectId + "'")
                    if len(chiAccounts) > 0:
                        chiAccount = chiAccounts.pop(0)
                        currency_id = _currencyIdConversion(self, chiAccount.CurrId)
                        account_type_id = self._accountTypeIdConversion(chiAccount.SubClsId, chiAccount.SubjectId)
                        reconcile = True if (account_type_id in (1, 2)) else False
                        parent_id = self._odooAccountIdFromCode(chiAccount.ParentSubId)
                        if not parent_id:
                            parent_id = None
                        accountValues = ({
                            'name': chiAccount.SubjectName,
                            'currency_id': currency_id,
                            'code': chiAccount.SubjectId,
                            'user_type_id': account_type_id,
                            'note': chiAccount.Description,
                            'reconcile': reconcile,
                            'x_batom_type_id': batomAccountTypeIds[chiAccount.SubClsId],
                            'x_batom_parent_id': parent_id,
                            })
                        
                        odooAccounts = accountModel.search([('code', '=', chiAccount.SubjectId)])
                        if len(odooAccounts) == 0:
                            accountModel.create(accountValues)
                        else:
                            odooAccounts[0].write(accountValues)
                except Exception:
                    _logger.warning('Exception in migrate_chart_of_account:', exc_info=True)
                    continue
            self.env.cr.commit()            
        except Exception:
            _logger.warning('Exception in migrate_chart_of_account:', exc_info=True)
    
class BatomMigrateProduct(models.TransientModel):
    _name = "batom.migrate_product"
    _description = "Migrate Products"
        
    def _migrate_chiProduct(self, cursorChi):
        # base_external_dbsource seems to have size limit of the query result
        # query the values directly may not returns all qualified records
        chiProductIds = cursorChi.execute('SELECT ProdId FROM comProduct ORDER BY ProdId').fetchall()
        productIds = []
        for chiProductId in chiProductIds:
            productIds.append(chiProductId.ProdId)
            
        nCount = len(productIds)
        nDone = 0
        for productId in productIds:
            try:
                _createOdooProduct(self, cursorChi, productId)
            except Exception:
                _logger.warning('Exception in migrate_product:', exc_info=True)
                import pdb; pdb.set_trace()
                continue
            nDone += 1
            if nDone % 10 == 0:
                print str(nDone) + '/' + str(nCount)
                self.env.cr.commit()
        self.env.cr.commit()
        
    def _migrate_inProduct(self, cursorBatom):
        inProducts = cursorBatom.execute('SELECT ProdId, ProdName, EngName, Remark, Unit FROM Product ORDER BY ProdId').fetchall()
        productModel = self.env['product.product']
        nCount = len(inProducts)
        nDone = 0
        for inProduct in inProducts:
            try:
                odooProducts = productModel.search([('default_code', '=', inProduct.ProdId)])
                if len(odooProducts) == 0:
                    if inProduct.ProdName and inProduct.ProdName.strip():
                        name = inProduct.ProdName
                    else:
                        name = inProduct.ProdId
                    uom_id = _uomIdConversion(self, inProduct.Unit.decode('utf-8'))
                    sale_ok = False
                    purchase_ok = False
                    type = 'product'
                    tracking = 'lot'
                    productValues = ({
                        'name': name,
                        'default_code': inProduct.ProdId,
                        'x_saved_code': inProduct.ProdId,
                        'type': type,
                        'tracking': tracking,
                        'sale_ok': sale_ok,
                        'purchase_ok': purchase_ok,
                        'description': inProduct.Remark,
                        'uom_id': uom_id,
                        'uom_po_id': uom_id,
                        })
                    
                    odooProduct = productModel.create(productValues)
                    if inProduct.EngName and inProduct.EngName.strip():
                        engName = inProduct.EngName
                        _updateTranslation(self, 'product.template,name', odooProduct.product_tmpl_id.id, engName, name)
            except Exception:
                _logger.warning('Exception in migrate_product:', exc_info=True)
                continue
            nDone += 1
            if nDone % 10 == 0:
                print str(nDone) + '/' + str(nCount)
                self.env.cr.commit()
        self.env.cr.commit()
        
    def _migrate_chiProcess(self, cursorChi):
        processes = cursorChi.execute('SELECT ProgramId, ProgramName, Remark FROM prdMakeProgram ORDER BY ProgramId').fetchall()
        productModel = self.env['product.product']
        nCount = len(processes)
        nDone = 0
        for process in processes:
            try:
                name = process.ProgramName
                sale_ok = False
                purchase_ok = True
                type = 'service'
                productValues = ({
                    'name': name,
                    'default_code': process.ProgramId,
                    'x_saved_code': process.ProgramId,
                    'type': type,
                    'sale_ok': sale_ok,
                    'purchase_ok': purchase_ok,
                    'x_is_process': True,
                    })
                
                odooProducts = productModel.search([('default_code', '=', process.ProgramId)])
                if len(odooProducts) == 0:
                    odooProduct = productModel.create(productValues)
                else:
                    odooProduct = odooProducts[0]
                    odooProduct.write(productValues)
                if process.Remark and process.Remark.strip():
                    engName = process.Remark
                    _updateTranslation(self, 'product.template,name', odooProduct.product_tmpl_id.id, engName, name)
            except Exception:
                _logger.warning('Exception in migrate_product:', exc_info=True)
                continue
            nDone += 1
            if nDone % 10 == 0:
                print str(nDone) + '/' + str(nCount)
                self.env.cr.commit()
        self.env.cr.commit()
        
    def _migrate_inProcess(self, cursorBatom):
        processes = cursorBatom.execute('SELECT ProcessId, ProcessName, Remark FROM Process ORDER BY ProcessId').fetchall()
        productModel = self.env['product.product']
        nCount = len(processes)
        nDone = 0
        for process in processes:
            try:
                if process.ProcessId and process.ProcessId.strip():
                    processId = process.ProcessId.strip()
                    odooProducts = productModel.search([('default_code', '=', processId)])
                    if len(odooProducts) == 0:
                        name = process.ProcessName
                        sale_ok = False
                        purchase_ok = True
                        type = 'service'
                        productValues = ({
                            'name': name,
                            'default_code': processId,
                            'type': type,
                            'sale_ok': sale_ok,
                            'purchase_ok': purchase_ok,
                            'description': process.Remark,
                            'x_is_process': True,
                            })
                        
                        odooProduct = productModel.create(productValues)
                        if process.Remark and process.Remark.strip():
                            engName = process.Remark
                            _updateTranslation(self, 'product.template,name', odooProduct.product_tmpl_id.id, engName, name)
            except Exception:
                _logger.warning('Exception in migrate_product:', exc_info=True)
                continue
            nDone += 1
            if nDone % 10 == 0:
                print str(nDone) + '/' + str(nCount)
                self.env.cr.commit()
        self.env.cr.commit()
        
    def _migrate_inShopProcess(self, cursorBatom):
        processes = cursorBatom.execute('SELECT ProcessId, ShopProcess, Type FROM ShopProcess ORDER BY ProcessId').fetchall()
        productModel = self.env['product.product']
        for process in processes:
            try:
                if process.ProcessId and process.ProcessId.strip():
                    processId = process.ProcessId.strip()
                    odooProducts = productModel.search([('default_code', '=', processId)])
                    if len(odooProducts) == 0:
                        name = process.ShopProcess
                        sale_ok = False
                        purchase_ok = True
                        type = 'service'
                        productValues = ({
                            'name': name,
                            'default_code': processId,
                            'type': type,
                            'sale_ok': sale_ok,
                            'purchase_ok': purchase_ok,
                            'description': process.Type,
                            'x_is_process': True,
                            })
                        
                        odooProduct = productModel.create(productValues)
            except Exception:
                _logger.warning('Exception in migrate_product:', exc_info=True)
                continue
        self.env.cr.commit()
            
    @api.multi
    def migrate_product(self):
        try:
            self.ensure_one()
            dbChi = self.env['base.external.dbsource'].search([('name', '=', 'CHIComp01')])
            dbBatom = self.env['base.external.dbsource'].search([('name', '=', 'Batom')])
            connChi = dbChi.conn_open()
            cursorChi = connChi.cursor()
            connBatom = dbBatom.conn_open()
            cursorBatom = connBatom.cursor()

            self._migrate_chiProduct(cursorChi)
            self._migrate_inProduct(cursorBatom)
            self._migrate_chiProcess(cursorChi)
            self._migrate_inProcess(cursorBatom)
            self._migrate_inShopProcess(cursorBatom)
            connChi.close()
            connBatom.close()
        except Exception:
            _logger.warning('Exception in migrate_product:', exc_info=True)
            if connChi:
                connChi.close()
            if connBatom:
                connBatom.close()
    
class BatomMigrateBom(models.TransientModel):
    _name = "batom.migrate_bom"
    _description = "Migrate BoM"
    _defaultMaterialAcquisitionProcess = None
    _defaultMaterialAcquisitionWorkcenter = None
    
    def _getDefaultMaterialAcquisitionProcess(self):
        process = None
        code = 'MTAC'
        name = '領料'
        try:
            processes = self.env['product.template'].search([
                ('x_is_process', '=', True),
                ('default_code', '=', code)
                ])
            if len(processes) > 0:
                process = processes[0]
            else:
                sale_ok = False
                purchase_ok = True
                type = 'service'
                process = self.env['product.product'].create({
                    'name': name,
                    'default_code': code,
                    'x_saved_code': code,
                    'type': type,
                    'sale_ok': sale_ok,
                    'purchase_ok': purchase_ok,
                    'x_is_process': True,
                    })
                _updateTranslation(self, 'product.template,name', process.product_tmpl_id.id, 'material acquisition', name)
        except Exception:
            _logger.warning('Exception in migrate_bom:', exc_info=True)
            
        return process
    
    def _getDefaultMaterialAcquisitionWorkcenter(self):
        if self._defaultMaterialAcquisitionProcess == None:
            self._defaultMaterialAcquisitionProcess = self._getDefaultMaterialAcquisitionProcess()
        return _getWorkcenter(self, self._defaultMaterialAcquisitionProcess.default_code, None, True)
    
    def _createRouting(self, chiBom, chiBomMaterials, chiBomProcesses, itemNo, active):
        odooRouting = None
        try:
            routingModel = self.env['mrp.routing']
            productTemplate = _getProductTemplate(self, chiBom.ProductId)
            if productTemplate == None:
                _logger.warning('Product template ' + chiBom.ProductId + ' not found in migrate_bom:', exc_info=True)
            else:
                odooRoutings = routingModel.search([
                    ('x_product_id', '=', productTemplate.id),
                    ('x_batom_bom_no', '=', itemNo),
                    ])
                if len(odooRoutings) > 0:
                    odooRouting = odooRoutings[0]
                else:
                    routingValues = ({
                        'code': u'~' + chiBom.ProductId.decode('utf-8') + u"#" + str(itemNo),
                        'name': u'RO/' + chiBom.ProductName.decode('utf-8') + u" #" + str(itemNo),
                        'x_product_id': productTemplate.id,
                        'x_batom_bom_no': itemNo,
                        'note': u'由正航 ' + chiBom.ProductId.decode('utf-8') + u'(' + chiBom.ProductName.decode('utf-8') + u') BoM自動產生的製程路徑',
                        'active': active,
                        })
                    odooRouting = routingModel.create(routingValues)
                    routingWorkcenterModel = self.env['mrp.routing.workcenter']
                    sequence = 0
                    if len(chiBomMaterials) > 0:
                        if self._defaultMaterialAcquisitionWorkcenter == None:
                            self._defaultMaterialAcquisitionWorkcenter = self._getDefaultMaterialAcquisitionWorkcenter()
                        odooWorkcenter = self._defaultMaterialAcquisitionWorkcenter
                        time_cycle_manual = 1
                        sequence += 5
                        routingWorkcenterValues = ({
                            'sequence': sequence,
                            'name': odooWorkcenter.x_process_id.name,
                            'workcenter_id': odooWorkcenter.id,
                            'routing_id': odooRouting.id,
                            'note': u'由正航 ' + chiBom.ProductId.decode('utf-8') + u'(' + chiBom.ProductName.decode('utf-8') + u') BoM自動產生的' + odooWorkcenter.x_process_id.name + u'作業',
                            'time_mode': 'manual',
                            'time_cycle_manual': time_cycle_manual,
                            'inspection_method': 'self',
                            })
                        routingWorkcenter = routingWorkcenterModel.create(routingWorkcenterValues)
                    for chiBomProcess in chiBomProcesses:
                        odooWorkcenter = _getWorkcenter(self, chiBomProcess.MkPgmId, chiBomProcess.Producer, True)
                        time_cycle_manual = (
                            1 if (chiBomProcess.WorkTimeOfBatch == None or chiBomProcess.WorkTimeOfBatch <= 0)
                            else chiBomProcess.WorkTimeOfBatch
                            )
                        sequence += 5
                        routingWorkcenterValues = ({
                            'sequence': sequence,
                            'name': odooWorkcenter.x_process_id.name,
                            'workcenter_id': odooWorkcenter.id,
                            'routing_id': odooRouting.id,
                            'note': u'由正航 ' + chiBom.ProductId.decode('utf-8') + u'(' + chiBom.ProductName.decode('utf-8') + u') BoM自動產生的' + odooWorkcenter.x_process_id.name + u'作業',
                            'time_mode': 'manual',
                            'time_cycle_manual': time_cycle_manual,
                            })
                        routingWorkcenter = routingWorkcenterModel.create(routingWorkcenterValues)
        except Exception:
            _logger.warning('Exception in migrate_bom:', exc_info=True)
            
        return odooRouting
    
    def _createBomLines(self, odooBom, odooRouting, chiBom, chiBomMaterials, chiBomProcesses):
        bomLineModel = self.env['mrp.bom.line']
        bomLineModel.search([('bom_id', '=', odooBom.id)]).unlink()
        routingWorkcenterModel = self.env['mrp.routing.workcenter']
        routingWorkcenters = routingWorkcenterModel.search([
            ('routing_id', '=', odooRouting.id)],
            order='sequence',
            )
        if self._defaultMaterialAcquisitionProcess == None:
            self._defaultMaterialAcquisitionProcess = self._getDefaultMaterialAcquisitionProcess()
        idxMaterials = 0
        idxProcesses = 0
        sequence = 0
        while idxMaterials < len(chiBomMaterials) or idxProcesses < len(chiBomProcesses):
            try:
                bomLineValues = None
                if sequence == 0 and len(chiBomMaterials) > 0:
                    # add the material acquisition process as the first process
                    product = self._defaultMaterialAcquisitionProcess
                    if product == None:
                        _logger.warning('_defaultMaterialAcquisitionProcess not found in migrate_bom:', exc_info=True)
                    else:
                        routing_id = None
                        operation_id = None
                        # first routingWorkcenter is the added material acquisition workcenter
                        if len(routingWorkcenters) > 0:
                            routing_id = odooRouting.id
                            operation_id = routingWorkcenters[0].id
                        sequence += 5
                        bomLineValues = ({
                            'bom_id': odooBom.id,
                            'sequence': sequence,
                            'product_id': product.id,
                            'product_qty': 1,
                            'product_uom_id': product.uom_id.id,
                            'routing_id': routing_id,
                            'operation_id': operation_id,
                            })
                elif (idxMaterials < len(chiBomMaterials) and
                        (idxProcesses >= len(chiBomProcesses) or 
                        chiBomMaterials[idxMaterials].SerNo <= chiBomProcesses[idxProcesses].SerNo
                        )):
                    product = _getProduct(self, chiBomMaterials[idxMaterials].SubProdId)
                    if product == None:
                        _logger.warning('Product template ' + chiBomMaterials[idxMaterials].SubProdId + ' not found in migrate_bom:', exc_info=True)
                    else:
                        routing_id = None
                        operation_id = None
                        # first routingWorkcenter is the added material acquisition workcenter
                        if len(routingWorkcenters) > 0:
                            routing_id = odooRouting.id
                            operation_id = routingWorkcenters[0].id
                        sequence += 5
                        bomLineValues = ({
                            'bom_id': odooBom.id,
                            'sequence': sequence,
                            'product_id': product.id,
                            'product_qty': chiBomMaterials[idxMaterials].QtyOfBatch,
                            'product_uom_id': product.uom_id.id,
                            'routing_id': routing_id,
                            'operation_id': operation_id,
                            })
                    idxMaterials += 1
                else:
                    product = _getProduct(self, chiBomProcesses[idxProcesses].MkPgmId)
                    if product == None:
                        _logger.warning('Product template ' + chiBomProcesses[idxProcesses].MkPgmId + ' not found in migrate_bom:', exc_info=True)
                    else:
                        routing_id = None
                        operation_id = None
                        # assuming routingWorkcenters got one to one correspondence with chiBomProcesses
                        # skipping the first routingWorkcenter as it is the added material acquisition workcenter
                        if idxProcesses + 1 < len(routingWorkcenters):
                            routing_id = odooRouting.id
                            operation_id = routingWorkcenters[idxProcesses + 1].id
                        sequence += 5
                        bomLineValues = ({
                            'bom_id': odooBom.id,
                            'sequence': sequence,
                            'product_id': product.id,
                            'product_qty': 1,
                            'product_uom_id': product.uom_id.id,
                            'routing_id': routing_id,
                            'operation_id': operation_id,
                            })
                    idxProcesses += 1
                if bomLineValues != None:
                    bomLineModel.create(bomLineValues)
            except Exception:
                _logger.warning('Exception in migrate_bom:', exc_info=True)
                continue
                
    def _migrate_chiBom(self, cursorChi):
        chiBoms = cursorChi.execute('SELECT ProductId, ProductName, ItemNo, CurVersion, Flag, NorProdtMode, BatchAmount, EffectDate FROM prdBOMMain ORDER BY ProductId, ItemNo').fetchall()
        chiBomVersions = cursorChi.execute('SELECT ProductId, count(*) as count FROM prdBOMMain GROUP BY ProductId ORDER BY ProductId').fetchall()
        
        bomVersions = {}
        for chiBomVersion in chiBomVersions:
            bomVersions[chiBomVersion.ProductId] = chiBomVersion.count
            
        nCount = len(chiBoms)
        nDone = 0
        for chiBom in chiBoms:
            try:
                if chiBom.ItemNo == 0:
                    itemNo = bomVersions[chiBom.ProductId]
                    active = True
                else:
                    itemNo = chiBom.ItemNo
                    active = False
                _createOdooBom(self, cursorChi, chiBom, itemNo, active)
            except Exception:
                _logger.warning('Exception in migrate_bom:', exc_info=True)
                continue
            nDone += 1
            if nDone % 10 == 0:
                print str(nDone) + '/' + str(nCount)
                self.env.cr.commit()
        self.env.cr.commit()

    def _createProcessPrice(self, production, chiProduction, chiWorkorder):
        productSupplierInfoModel = self.env['product.supplierinfo']
                    
        try:
            supplier = _getSupplier(self, chiWorkorder.Producer, True)
            processProduct = _getProduct(self, chiWorkorder.MkPgmId)
            if supplier and processProduct:
                date_start = datetime.strptime(str(chiProduction.MkOrdDate), '%Y%m%d')
                productSupplierInfos = productSupplierInfoModel.search([
                    ('name', '=', supplier.id),
                    ('target_product_id', '=', production.product_id.id),
                    ('product_id', '=', processProduct.id),
                    ], order='date_start')
                newPrice = True
                # the following logic will merge the same price entries into date ranges
                # it does not try to handle all cases perfectly
                if len(productSupplierInfos) > 0:
                    i = 0
                    idxToInsert = -1
                    previousDate = False
                    previousDateSamePrice = False
                    nextDate = False
                    nextDateSamePrice = False
                    while i < len(productSupplierInfos):
                        productSupplierInfo = productSupplierInfos[i]
                        supplierInfoDateStart = datetime.strptime(productSupplierInfo.date_start, '%Y-%m-%d')
                        if date_start <= supplierInfoDateStart:
                            if idxToInsert < 0:
                                idxToInsert = i
                                nextDate = supplierInfoDateStart
                                if chiWorkorder.Price == productSupplierInfo.price:
                                    nextDateSamePrice = True
                                    break
                            if date_start == supplierInfoDateStart:
                                if chiWorkorder.Price == productSupplierInfo.price:
                                    newPrice = False # duplicate. flag it as no action needed
                                    break
                        elif idxToInsert >= 0:
                            if nextDate == supplierInfoDateStart:
                                if chiWorkorder.Price == productSupplierInfo.price:
                                    nextDateSamePrice = True
                                    break
                            else:
                                break
                        if idxToInsert < 0 and previousDate != supplierInfoDateStart:
                            previousDate = supplierInfoDateStart
                            if chiWorkorder.Price == productSupplierInfo.price:
                                previousDateSamePrice = True
                            else:
                                previousDateSamePrice = False
                        i += 1
                    
                    if newPrice:
                        if idxToInsert < 0:
                            idxToInsert = i
                        if previousDateSamePrice and nextDateSamePrice:
                            newPrice = False # inside an existing date range with the same price.  no action needed
                        elif previousDateSamePrice:
                            if (idxToInsert >= 2 and
                                    productSupplierInfos[idxToInsert - 1].price == chiWorkorder.Price and
                                    productSupplierInfos[idxToInsert - 2].price == chiWorkorder.Price):
                                productSupplierInfos[idxToInsert - 1].write({
                                    'date_start': date_start,
                                    })
                                newPrice = False
                        elif nextDateSamePrice:
                            if (idxToInsert < len(productSupplierInfos) - 1 and
                                    productSupplierInfos[idxToInsert].price == chiWorkorder.Price and
                                    productSupplierInfos[idxToInsert + 1].price == chiWorkorder.Price):
                                productSupplierInfos[idxToInsert].write({
                                    'date_start': date_start,
                                    })
                                newPrice = False
                        # else: create new price
                        
                if newPrice:
                    productSupplierInfoValues = ({
                        'name': supplier.id,
                        'product_name': supplier.display_name + ' -> ' + processProduct.name,
                        'product_code': processProduct.default_code + u'->' + production.product_id.default_code,
                        'price': chiWorkorder.Price,
                        'date_start': date_start,
                        'product_id': processProduct.id,
                        'product_tmpl_id': processProduct.product_tmpl_id.id,
                        'target_product_id': production.product_id.id
                        })
                    productSupplierInfoModel.create(productSupplierInfoValues)
        except Exception:
            _logger.warning('Exception in migrate_bom:', exc_info=True)
            import pdb; pdb.set_trace()
        
    def _migrate_inRouting(self, cursorBatom):
        inProducts = cursorBatom.execute('SELECT ProdId, ProdName, EngName, Remark, Unit FROM Product ORDER BY ProdId').fetchall()
        productModel = self.env['product.product']
        for inProduct in inProducts:
            try:
                odooProducts = productModel.search([('default_code', '=', inProduct.ProdId)])
                if len(odooProducts) == 0:
                    if inProduct.ProdName and inProduct.ProdName.strip():
                        name = inProduct.ProdName
                    else:
                        name = inProduct.ProdId
                    uom_id = _uomIdConversion(self, inProduct.Unit.decode('utf-8'))
                    sale_ok = False
                    purchase_ok = False
                    type = 'product'
                    productValues = ({
                        'name': name,
                        'default_code': inProduct.ProdId,
                        'type': type,
                        'sale_ok': sale_ok,
                        'purchase_ok': purchase_ok,
                        'description': inProduct.Remark,
                        'uom_id': uom_id,
                        'uom_po_id': uom_id,
                        })
                    
                    odooProduct = productModel.create(productValues)
                    if inProduct.EngName and inProduct.EngName.strip():
                        engName = inProduct.EngName
                        _updateTranslation(self, 'product.template,name', odooProduct.product_tmpl_id.id, engName, name)
            except Exception:
                _logger.warning('Exception in migrate_product:', exc_info=True)
                continue
        self.env.cr.commit()
            
    @api.multi
    def migrate_bom(self):
        try:
            self.ensure_one()
            dbChi = self.env['base.external.dbsource'].search([('name', '=', 'CHIComp01')])
            dbBatom = self.env['base.external.dbsource'].search([('name', '=', 'Batom')])
            connChi = dbChi.conn_open()
            cursorChi = connChi.cursor()
            connBatom = dbBatom.conn_open()
            cursorBatom = connBatom.cursor()

            self._migrate_chiBom(cursorChi)
            connChi.close()
            connBatom.close()
        except Exception:
            _logger.warning('Exception in migrate_bom:', exc_info=True)
            if connChi:
                connChi.close()
            if connBatom:
                connBatom.close()
    
    def _matchBom(self, odooBom, materials, workorders):
        try:
            matched = True
            odooBomLines = odooBom.bom_line_ids
            if odooBomLines[0].product_id.default_code == 'MTAC':
                odooBomLines = odooBomLines[1:]
            for material in materials:
                i = 0
                found = False
                while i < len(odooBomLines):
                    if (not odooBomLines[i].product_id.product_tmpl_id.x_is_process and
                            odooBomLines[i].product_id.default_code == material.SubProdID and
                            odooBomLines[i].product_qty == material.UnitOughtQty):
                        odooBomLines = odooBomLines[:i] + odooBomLines[i + 1:]
                        found = True
                        break
                    else:
                        i += 1
                if not found:
                    matched = False
                    break
                    
            if matched:
                for workorder in workorders:
                    if workorder.Producer == '01':
                        workorder.Producer = False
                    if (odooBomLines and
                            odooBomLines[0].operation_id.workcenter_id.x_process_id.default_code == workorder.MkPgmID and
                            odooBomLines[0].operation_id.workcenter_id.x_supplier_id.x_supplier_code == workorder.Producer):
                        odooBomLines = odooBomLines[1:]
                    else:
                        matched = False
                        break
        except Exception:
            _logger.warning('Exception in _matchBom:', exc_info=True)
            import pdb; pdb.set_trace()
        
        return matched
        
    def _getBom(self, cursorChi, chiMo):
        matchedOdooBom = False
        templates = self.env['product.template'].search([('default_code', '=', chiMo.ProductId)])
        if templates:
            template = templates[0]
        else:
            odooProduct = _createOdooProduct(self, cursorChi, productId)
            template = odooProduct.product_tmpl_id
            
        if template:
            odooBoms = self.env['mrp.bom'].search([
                '|',
                ('active', '=', True),
                ('active', '=', False),
                ('product_tmpl_id', '=', template.id),
                ])
            if not odooBoms:
                chiBoms = cursorChi.execute(
                    u"SELECT ProductId, ProductName, ItemNo, CurVersion, Flag, NorProdtMode, BatchAmount, EffectDate "
                    u"FROM prdBOMMain WHERE ProductId ='" + chiMo.ProductId.decode('utf-8') + u"' "
                    u"ORDER ItemNo").fetchall()
                for chiBom in chiBoms:
                    try:
                        if chiBom.ItemNo == 0:
                            itemNo = len(chiBoms)
                            active = True
                        else:
                            itemNo = chiBom.ItemNo
                            active = False
                        odooBoms.append(_createOdooBom(self, cursorChi, chiBom, itemNo, active))
                    except Exception:
                        _logger.warning('Exception in migrate_bom:', exc_info=True)
                        import pdb; pdb.set_trace()
                        continue
            if odooBoms:
                materials = cursorChi.execute(
                    "SELECT SerNo, SubProdID, ProcRowNO, UnitOughtQty "
                    "FROM prdMkOrdMats "
                    "WHERE MkOrdNo='" + chiMo.MkOrdNo + "' "
                    "ORDER BY SerNo").fetchall()
                workorders = cursorChi.execute(
                    "SELECT SerNo, RowNO, MkPgmID, Producer, Price "
                    "FROM prdMkOrdPgms "
                    "WHERE MkOrdNo='" + chiMo.MkOrdNo + "' "
                    "ORDER BY SerNo").fetchall()
                for odooBom in odooBoms:
                    if self._matchBom(odooBom, materials, workorders):
                        matchedOdooBom = odooBom
                        break
                if not matchedOdooBom:
                    matchedOdooBom = odooBoms[0]
        return matchedOdooBom

    def _migrate_chiMo(self, cursorChi, cursorBatom, mo_id, flag):
        productionModel = self.env['mrp.production']
        workorderModel = self.env['mrp.workorder']

        if flag == 1:
            name = 'M' + mo_id[1:]
        elif flag == 2:
            name = 'N' + mo_id[1:]
        else:
            name = mo_id
        production = productionModel.search([('name', '=', name)])
        if not production:
            sql = (
                "SELECT MkOrdNo, MkOrdDate, ProductId, ProdtQty, MakerId "
                "FROM prdMkOrdMain "
                "WHERE MkOrdNo='" + mo_id + "' AND Flag=" + str(flag))
            manufactureOrders = cursorChi.execute(sql).fetchall()
            
            if manufactureOrders:
                try:
                    manufactureOrder = manufactureOrders[0]
                    targetBom = self._getBom(cursorChi, manufactureOrder)
                    if targetBom:
                        products = self.env['product.product'].search([
                            ('default_code', '=', manufactureOrder.ProductId)
                            ])
                        product = products[0] if products else False
                        users = self.env['res.users'].search([
                            ('login', '=', manufactureOrder.MakerId)
                            ])
                        user = users[0] if users else False
                        productionValues = ({
                            'name': name,
                            'product_id': product.id if product else False,
                            'bom_id': targetBom.id,
                            'routing_id': targetBom.routing_id.id,
                            'product_qty': manufactureOrder.ProdtQty,
                            'product_uom_id': product.product_tmpl_id.uom_id.id if product else False,
                            'user_id': user.id if user else False,
                            })
                        production = productionModel.create(productionValues)
                        production.button_plan()
                        
                        sql = (
                            "SELECT MkPgmId, Producer, Price "
                            "FROM prdMkOrdPgms "
                            "WHERE MkOrdNo='" + mo_id + "' AND Flag=" + str(flag) + " "
                            "ORDER BY SerNo")
                        workorders = cursorChi.execute(sql).fetchall()
                        for workorder in workorders:
                            if workorder.Price > 0:
                                self._createProcessPrice(production, manufactureOrder, workorder)
                except Exception:
                    _logger.warning('Exception in migrate_bom:', exc_info=True)
                    import pdb; pdb.set_trace()
            self.env.cr.commit()
        return production
            
    @api.multi
    def migrate_mo(self, mo_id, flag = 3):
        try:
            self.ensure_one()
            dbChi = self.env['base.external.dbsource'].search([('name', '=', 'CHIComp01')])
            dbBatom = self.env['base.external.dbsource'].search([('name', '=', 'Batom')])
            connChi = dbChi.conn_open()
            cursorChi = connChi.cursor()
            connBatom = dbBatom.conn_open()
            cursorBatom = connBatom.cursor()

            self._migrate_chiMo(cursorChi, cursorBatom, mo_id, flag)
            connChi.close()
            connBatom.close()
        except Exception:
            _logger.warning('Exception in migrate_mo:', exc_info=True)
            import pdb; pdb.set_trace()
            if connChi:
                connChi.close()
            if connBatom:
                connBatom.close()
    
    def _unmergeProductAndLotCells(self, sheet):
        for range_ in sheet.merged_cell_ranges:
            merged_cells = list(rows_from_range(range_))
            rows = len(merged_cells)
            xy = coordinate_from_string(merged_cells[0][0])
            if rows > 1 and (xy[0] == 'A' or xy[0] == 'I'): # product cod column or lot code column
                sheet.unmerge_cells(merged_cells[0][0] + ':' + merged_cells[rows - 1][0])
                for i in xrange(1, rows):
                    sheet[merged_cells[i][0]].value = sheet[merged_cells[0][0]].value

    def _isQuantity(self, cell):
        return cell.value and cell.data_type == 'n' # number cell
    
    def _getMaterialLot(self, row):
        lot = False
        if row[8].value:
            if row[8].data_type == 's': # text cell
                lot = row[8].value.strip()
            else:
                lot = str(row[8].value)
            if (not lot or # blank
                    lot.upper() == 'X' or
                    lot == u'待通知'):
                lot = False
        return lot
    
    def _getCost(self, cell):
        cost = False
        if self._isQuantity(cell):
            cost = cell.value
        elif cell.value:
            values = cell.value.strip().split()
            for value in values:
                try:
                    cost = float(value)
                    break
                except ValueError:
                    pass
        return cost
    
    def _getCustCost(self, row):
        return self._getCost(row[1])
        
    def _getMaterialCost(self, row):
        return self._getCost(row[2])
    
    def _getForgeCost(self, row):
        return self._getCost(row[4])
    
    def _getAdditionalRowsInMergedCells(self, sheet, cell):
        additionalRowsInMergedCells = 0
        idx = cell.coordinate
        for range_ in sheet.merged_cell_ranges:
            merged_cells = list(rows_from_range(range_))
            for row in merged_cells:
                if idx in row:
                    dditionalRowsInMergedCells = len(merged_cells) - 1
                    break
        return dditionalRowsInMergedCells
        
    def _getDispatchDate(self, row):
        dispatch_date = False
        try:
            if row[7].value and row[7].data_type == 's': # text cell
                value = row[7].value.strip()
                if len(value) > 4 and value[:4] == u'製表日期':
                    dateTokens = re.findall(r"[\w]+", value) # 製表日期 will be stripped out
                    if len(dateTokens) == 3:
                        dispatch_date = datetime(int(dateTokens[0]), int(dateTokens[1]), int(dateTokens[2])).date()
        except:
            pass
        return dispatch_date
        
    def _matchProductCode(self, productCodeCellValue, productCode):
        matched = False
        productCodeValues = re.findall(r"[\w-]+", productCodeCellValue)
        for productCodeValue in productCodeValues:
            if productCode.find(productCodeValue) >= 0:
                matched = True
                break
        return matched
    
    def _updateMaterialAssignmentCodes(self, cursorChi, materialAssignment, productCodeCellValue, row):
        try:
            if row[7].value:
                code = row[7].value.strip().upper()
                if code:
                    adjustment_code = False
                    inventory_flag = False
                    inventory_code = False
                    production_code = False
                    results = False
                    if code[:1] == u'調' and code[1:2] == 'A':
                        adjustment_code = code[1:]
                        results = cursorChi.execute(
                            "SELECT ProdId, Quantity "
                            "FROM comProdRec "
                            "WHERE Flag = '300' AND BillNO = '" + adjustment_code + "'"
                            ).fetchall()
                    elif code[:1] == u'進' and code[1:2] == 'A':
                        inventory_flag = '100'
                        inventory_code = code[1:]
                        results = cursorChi.execute(
                            "SELECT ProdId, Quantity "
                            "FROM comProdRec "
                            "WHERE Flag = '100' AND BillNO = '" + inventory_code + "'"
                            ).fetchall()
                    elif code[:1] == 'A':
                        production_code = code
                        results = cursorChi.execute(
                            "SELECT comProdRec.Flag, BillNo, ProdId, Quantity "
                            "FROM comProdRec INNER JOIN "
                            "(SELECT Flag, WareInNo FROM prdWareIn "
                            "WHERE Flag IN ('311', '312') AND "
                            "MkOrdNo = '" + production_code + "') tmpWareIn ON "
                            "(comProdRec.Flag = tmpWareIn.Flag AND comProdRec.BillNo = tmpWareIn.WareInNo)"
                            ).fetchall()
                    if results:
                        matched = False
                        message = ""
                        for result in results:
                            if self._matchProductCode(productCodeCellValue, result.ProdId):
                                matched = True
                                if not materialAssignment.product_code:
                                    materialAssignment.product_code = result.ProdId
                                    materialAssignment.product_qty = result.Quantity
                                    materialAssignment.adjustment_code = adjustment_code
                                    materialAssignment.inventory_code = inventory_code
                                    materialAssignment.production_code = production_code
                                elif materialAssignment.product_code == result.ProdId:
                                    materialAssignment.product_qty += result.Quantity
                            else:
                                message =  'product code "' + productCodeCellValue + '" does not match "' + result.ProdId + '" "' + row[7].value + '"'
                        if not matched:
                            print message
        except:
            _logger.warning('Exception in _updateMaterialAssignmentCodes:', exc_info=True)
            
    def _createMaterialAssignments(self, materialAssignments):
        materialAssignmentModel = self.env['batom.material_assignment']
        for materialAssignment in materialAssignments:
            if materialAssignment.production_code or materialAssignment.inventory_code or materialAssignment.adjustment_code:
                if (materialAssignmentModel.search([
                        ('production_code', '=', materialAssignment.production_code),
                        ('inventory_code', '=', materialAssignment.inventory_code),
                        ('adjustment_code', '=', materialAssignment.adjustment_code),
                        ])):
                    print ('Material assignment for' +
                        (' MO ' + materialAssignment.production_code if materialAssignment.production_code else '') +
                        (' iventory in ' + materialAssignment.inventory_code if materialAssignment.inventory_code else '') +
                        (' iventory adjustoment ' + materialAssignment.adjustment_code if materialAssignment.adjustment_code else '') +
                        ' already exists')
                    materialAssignment = False
            else:
                materialAssignment = False
            if materialAssignment:
                materialAssignmentModel.create({
                    'product_code': materialAssignment.product_code,
                    'product_qty': materialAssignment.product_qty,
                    'production_code': materialAssignment.production_code,
                    'inventory_code': materialAssignment.inventory_code,
                    'adjustment_code': materialAssignment.adjustment_code,
                    'cut_cost': materialAssignment.cut_cost,
                    'material_cost': materialAssignment.material_cost,
                    'forge_cost': materialAssignment.forge_cost,
                    'material_lot': materialAssignment.material_lot,
                    'dispatch_date': materialAssignment.dispatch_date,
                    })
                print ('Creating material assignment for' +
                    (' MO ' + materialAssignment.production_code if materialAssignment.production_code else '') +
                    (' iventory in ' + materialAssignment.inventory_code if materialAssignment.inventory_code else '') +
                    (' iventory adjustoment ' + materialAssignment.adjustment_code if materialAssignment.adjustment_code else ''))

    def _getInventoryInFromChiMo(self, cursorChi, production_id, product_code, product_qty, date_start, date_end):
        returnedMaterialAssignments = []
        materialAssignments = self.env['batom.material_assignment'].search([
            ('product_code', '=', product_code),
            ('production_code', '!=', False),
            ('dispatch_date', '>=', date_start), 
            ('dispatch_date', '<=', date_end), 
            ('used', '=', False),
            ], order='dispatch_date desc')
        found = False
        inventoryMaterialAssignments = []
        for materialAssignment in materialAssignments:
            inventoryIns = cursorChi.execute(
                "SELECT comProdRec.Flag, BillNo, ProdId, Quantity "
                "FROM comProdRec INNER JOIN "
                "(SELECT Flag, WareInNo FROM prdWareIn "
                "WHERE Flag IN ('311', '312') AND "
                "MkOrdNo = '" + materialAssignment.production_code + "') tmpWareIn ON "
                "(comProdRec.Flag = tmpWareIn.Flag AND comProdRec.BillNo = tmpWareIn.WareInNo)"
                ).fetchall()
            for inventoryIn in inventoryIns:
                justFound = False
                if not found and inventoryIn.Quantity == product_qty:
                    found = True
                    justFound = True
                inventoryMaterialAssignment = self.env['batom.material_assignment'].create({
                    'product_code': product_code,
                    'product_qty': product_qty,
                    'production_code': False,
                    'inventory_flag': str(inventoryIn.Flag),
                    'inventory_code': inventoryIn.BillNo,
                    'adjustment_code': False,
                    'cut_cost': materialAssignment.cut_cost,
                    'material_cost': materialAssignment.material_cost,
                    'forge_cost': materialAssignment.forge_cost,
                    'material_lot': materialAssignment.material_lot,
                    'dispatch_date': materialAssignment.dispatch_date,
                    'used': justFound,
                    'used_by_production_ids': [(4, [production_id])] if justFound else False,
                    })
                if justFound:
                    returnedMaterialAssignments.append(inventoryMaterialAssignment)
            if found:
                self.env.cr.commit()
                break
        return returnedMaterialAssignments
                
    @api.multi
    def migrate_mo_by_date(self, dateStart, dateEnd = '99999999'):
        self.ensure_one()
        if len(dateStart) != 8 or not dateStart.isdigit():
            print '"' + dateStart + '" is not a valid date'
            return
        if len(dateEnd) != 8 or not dateEnd.isdigit():
            print '"' + dateEnd + '" is not a valid date'
            return
            
        try:
            dbChi = self.env['base.external.dbsource'].search([('name', '=', 'CHIComp01')])
            dbBatom = self.env['base.external.dbsource'].search([('name', '=', 'Batom')])
            connChi = dbChi.conn_open()
            cursorChi = connChi.cursor()
            connBatom = dbBatom.conn_open()
            cursorBatom = connBatom.cursor()
            
            sql = (
                "SELECT Flag, MkOrdNo "
                "FROM prdMkOrdMain "
                "WHERE MkOrdDate BETWEEN " + dateStart + " AND " + dateEnd + " AND Flag in (1, 2, 3)")
            manufactureOrders = cursorChi.execute(sql).fetchall()
            nCount = len(manufactureOrders)
            nDone = 0
            for manufactureOrder in manufactureOrders:
                self._migrate_chiMo(cursorChi, cursorBatom, manufactureOrder.MkOrdNo, manufactureOrder.Flag)
                nDone += 1
                if nDone % 10 == 0:
                    print str(nDone) + '/' + str(nCount)
                    self.env.cr.commit()
            self.env.cr.commit()
            connChi.close()
            connBatom.close()
        except Exception:
            _logger.warning('Exception in migrate_mo_by_date:', exc_info=True)
            import pdb; pdb.set_trace()
            if connChi:
                connChi.close()
            if connBatom:
                connBatom.close()
            
    @api.multi
    def load_lot_sheet(self, xlsx_file):
        self.ensure_one()
        try:
            dbChi = self.env['base.external.dbsource'].search([('name', '=', 'CHIComp01')])
            connChi = dbChi.conn_open()
            cursorChi = connChi.cursor()
            materialAssignments = []
            nMaterialAssignments = 0
            idxToAssignDispatchDate = 0
            wb = load_workbook(filename = xlsx_file)
            ws = wb.active
            self._unmergeProductAndLotCells(ws)
            mergedCells = ws.merged_cells
            additionalRowsInMergedCells = 0
            materialAssignment = False
            productCodeCellValue = False
            for row in ws.iter_rows():
                if additionalRowsInMergedCells == 0:
                    if materialAssignment:
                        if materialAssignment.product_code:
                            materialAssignments.append(materialAssignment)
                            nMaterialAssignments += 1
                        materialAssignment = False
                    if self._isQuantity(row[3]):
                        if row[0].value:
                            productCodeCellValue = row[0].value
                        material_lot = self._getMaterialLot(row)
                        if not material_lot:
                            materialAssignment = False
                        else:
                            materialAssignment = _material_assignment(
                                product_code = False,
                                product_qty = False,
                                production_code = False,
                                inventory_flag = False,
                                inventory_code = False,
                                adjustment_code = False,
                                cut_cost = self._getCustCost(row),
                                material_cost = self._getMaterialCost(row),
                                forge_cost = self._getForgeCost(row),
                                material_lot = material_lot,
                                dispatch_date = False)
                            if row[3].coordinate in mergedCells:
                                additionalRowsInMergedCells = self._getAdditionalRowsInMergedCells(ws, row[3])
                            self._updateMaterialAssignmentCodes(cursorChi, materialAssignment, productCodeCellValue, row)
                    elif idxToAssignDispatchDate < nMaterialAssignments:
                        dispatch_date = self._getDispatchDate(row)
                        if dispatch_date:
                            while idxToAssignDispatchDate < nMaterialAssignments:
                                materialAssignments[idxToAssignDispatchDate].dispatch_date = dispatch_date
                                idxToAssignDispatchDate += 1
                else:
                    additionalRowsInMergedCells -= 1
                    if materialAssignment and not materialAssignment.product_code:
                        self._updateMaterialAssignmentCodes(cursorChi, materialAssignment, productCodeCellValue, row)

            self._createMaterialAssignments(materialAssignments)
            self.env.cr.commit()
            connChi.close()
        except Exception:
            _logger.warning('Exception in load_lot_sheet:', exc_info=True)
            import pdb; pdb.set_trace()
            if connChi:
                connChi.close()

    def _check_migrate_workorders(self, cursorChi, cursorBatom, mo_id, flag):
        productionModel = self.env['mrp.production']
        workorderModel = self.env['mrp.workorder']

        if flag == 1:
            productionCode = 'M' + mo_id[1:]
        elif flag == 2:
            productionCode = 'N' + mo_id[1:]
        else:
            productionCode = mo_id
        try:
            odooProductions = productionModel.search([('name', '=', productionCode)])
            if odooProductions:
                odooProduction = odooProductions[0]
            else:
                odooProduction = self._migrate_chiMo(cursorChi, cursorBatom, mo_id, flag)
            if odooProduction:
                materialAssignments = self.env['batom.material_assignment'].search([
                    ('used_by_production_ids', '=', odooProduction.id)
                    ])
                if materialAssignments:
                    for materialAssignment in materialAssignments:
                        if materialAssignment.production_code:
                            print ('MO ' + materialAssignment.production_code + ' is used for ' + productionCode + '/' +
                                materialAssignment.product_code + '/' +
                                str(materialAssignment.product_qty))
                        elif materialAssignment.inventory_code:
                            print ('Warehouse in ' + materialAssignment.inventory_code + ' is used for ' + productionCode + '/' +
                                materialAssignment.product_code + '/' +
                                str(materialAssignment.product_qty))
                        else:
                            print ('Inventory adjustment ' + materialAssignment.adjustment_code + ' is used for ' + productionCode + '/' +
                                materialAssignment.product_code + '/' +
                                str(materialAssignment.product_qty))
                else:
                    yyyy = int(productionCode[1:5])
                    mm = int(productionCode[5:7])
                    dd = int(productionCode[7:9])
                    date_end = datetime(yyyy, mm, dd).date()
                    date_start = date_end - timedelta(days = 180)
                    date_end = date_end + timedelta(days = 30)
                    bom_product_quants = self.env['mrp.bom.line'].search([
                        ('bom_id', '=', odooProduction.bom_id.id),
                        ('product_id.type', '=', 'product')
                        ])
                    for bom_product_quant in bom_product_quants:
                        materialAssignments = self.env['batom.material_assignment'].search([
                            '|',
                            ('used_by_production_ids', '=', odooProduction.id),
                            ('product_code', '=', bom_product_quant.product_id.default_code),
                            ('product_qty', '=', odooProduction.product_qty * bom_product_quant.product_qty),
                            ('dispatch_date', '>=', date_start), 
                            ('dispatch_date', '<=', date_end), 
                            ('used', '=', False),
                            ], order='dispatch_date desc')
                        if not materialAssignments:
                            materialAssignments = self._getInventoryInFromChiMo(cursorChi, odooProduction.id, bom_product_quant.product_id.default_code, odooProduction.product_qty * bom_product_quant.product_qty, date_start, date_end)
                        if not materialAssignments:
                            print ('Corresponding material assignment could not be fond for ' + productionCode + '/' +
                                bom_product_quant.product_id.default_code + '/' +
                                str(odooProduction.product_qty * bom_product_quant.product_qty))
                        else:
                            materialAssignment = materialAssignments[0]
                            if odooProduction not in materialAssignment.used_by_production_ids:
                                materialAssignment.write({
                                    'used_by_production_ids': [(4, [odooProduction.id])],
                                    'used': True,
                                    })
                            if materialAssignment.production_code:
                                print ('MO ' + materialAssignment.production_code + ' is used for ' + productionCode + '/' +
                                    bom_product_quant.product_id.default_code + '/' +
                                    str(odooProduction.product_qty * bom_product_quant.product_qty))
                            elif materialAssignment.inventory_code:
                                print ('Inventory in ' + 
                                    (materialAssignment.inventory_flag + '/' if materialAssignment.inventory_flag else '') +
                                    materialAssignment.inventory_code + ' is used for ' + productionCode + '/' +
                                    bom_product_quant.product_id.default_code + '/' +
                                    str(odooProduction.product_qty * bom_product_quant.product_qty))
                            else:
                                print ('Inventory adjustment ' + materialAssignment.adjustment_code + ' is used for ' + productionCode + '/' +
                                    bom_product_quant.product_id.default_code + '/' +
                                    str(odooProduction.product_qty * bom_product_quant.product_qty))
            else:
                print 'Manufacture order ' + productionCode + ' can not be found'
        except Exception:
            _logger.warning('Exception in _check_migrate_workorders:', exc_info=True)
            import pdb; pdb.set_trace()
            
    @api.multi
    def check_migrate_workorders(self, mo_id, flag = 3):
        try:
            self.ensure_one()
            dbChi = self.env['base.external.dbsource'].search([('name', '=', 'CHIComp01')])
            dbBatom = self.env['base.external.dbsource'].search([('name', '=', 'Batom')])
            connChi = dbChi.conn_open()
            cursorChi = connChi.cursor()
            connBatom = dbBatom.conn_open()
            cursorBatom = connBatom.cursor()

            self._check_migrate_workorders(cursorChi, cursorBatom, mo_id, flag)
            self.env.cr.commit()
            connChi.close()
            connBatom.close()
        except Exception:
            _logger.warning('Exception in check_migrate_workorders:', exc_info=True)
            import pdb; pdb.set_trace()
            if connChi:
                connChi.close()
            if connBatom:
                connBatom.close()
    
    @api.multi
    def check_migrate_workorders_by_date(self, dateStart, dateEnd = '99999999'):
        self.ensure_one()
        if len(dateStart) != 8 or not dateStart.isdigit():
            print '"' + dateStart + '" is not a valid date'
            return
        if len(dateEnd) != 8 or not dateEnd.isdigit():
            print '"' + dateEnd + '" is not a valid date'
            return
            
        try:
            dbChi = self.env['base.external.dbsource'].search([('name', '=', 'CHIComp01')])
            dbBatom = self.env['base.external.dbsource'].search([('name', '=', 'Batom')])
            connChi = dbChi.conn_open()
            cursorChi = connChi.cursor()
            connBatom = dbBatom.conn_open()
            cursorBatom = connBatom.cursor()
            
            sql = (
                "SELECT Flag, MkOrdNo "
                "FROM prdMkOrdMain "
                "WHERE MkOrdDate BETWEEN " + dateStart + " AND " + dateEnd + " AND Flag in (1, 2, 3)")
            manufactureOrders = cursorChi.execute(sql).fetchall()
            nCount = len(manufactureOrders)
            nDone = 0
            for manufactureOrder in manufactureOrders:
                self._check_migrate_workorders(cursorChi, cursorBatom, manufactureOrder.MkOrdNo, manufactureOrder.Flag)
                nDone += 1
                if nDone % 10 == 0:
                    print str(nDone) + '/' + str(nCount)
                    self.env.cr.commit()
            self.env.cr.commit()
            connChi.close()
            connBatom.close()
        except Exception:
            _logger.warning('Exception in check_migrate_workorders_by_date:', exc_info=True)
            import pdb; pdb.set_trace()
            if connChi:
                connChi.close()
            if connBatom:
                connBatom.close()

    def _migrate_inWorkorders(self, cursorChi, cursorBatom, mo_id, flag):
        productionModel = self.env['mrp.production']
        workorderModel = self.env['mrp.workorder']

        if flag == 1:
            productionCode = 'M' + mo_id[1:]
        elif flag == 2:
            productionCode = 'N' + mo_id[1:]
        else:
            productionCode = mo_id
        try:
            odooProductions = productionModel.search([('name', '=', productionCode)])
            if odooProductions:
                odooProduction = odooProductions[0]
            else:
                odooProduction = self._migrate_chiMo(cursorChi, cursorBatom, mo_id, flag)
            if odooProduction:
                self._check_migrate_workorders(cursorChi, cursorBatom, mo_id, flag)
                materialAssignments = self.env['batom.material_assignment'].search([
                    ('used_by_production_ids', '=', odooProduction.id)
                    ])
                #if materialAssignments:
                #    for materialAssignment in materialAssignments:
                #else:
            else:
                print 'Manufacture order ' + productionCode + ' can not be found'
        except Exception:
            _logger.warning('Exception in _migrate_inWorkorders:', exc_info=True)
            import pdb; pdb.set_trace()
            
    @api.multi
    def migrate_workorders(self, mo_id, flag = 3):
        try:
            self.ensure_one()
            dbChi = self.env['base.external.dbsource'].search([('name', '=', 'CHIComp01')])
            dbBatom = self.env['base.external.dbsource'].search([('name', '=', 'Batom')])
            connChi = dbChi.conn_open()
            cursorChi = connChi.cursor()
            connBatom = dbBatom.conn_open()
            cursorBatom = connBatom.cursor()

            self._migrate_inWorkorders(cursorChi, cursorBatom, mo_id, flag)
            self.env.cr.commit()
            connChi.close()
            connBatom.close()
        except Exception:
            _logger.warning('Exception in migrate_workorders:', exc_info=True)
            import pdb; pdb.set_trace()
            if connChi:
                connChi.close()
            if connBatom:
                connBatom.close()
    
    @api.multi
    def migrate_workorders_by_date(self, dateStart, dateEnd = '99999999'):
        self.ensure_one()
        if len(dateStart) != 8 or not dateStart.isdigit():
            print '"' + dateStart + '" is not a valid date'
            return
        if len(dateEnd) != 8 or not dateEnd.isdigit():
            print '"' + dateEnd + '" is not a valid date'
            return
            
        try:
            dbChi = self.env['base.external.dbsource'].search([('name', '=', 'CHIComp01')])
            dbBatom = self.env['base.external.dbsource'].search([('name', '=', 'Batom')])
            connChi = dbChi.conn_open()
            cursorChi = connChi.cursor()
            connBatom = dbBatom.conn_open()
            cursorBatom = connBatom.cursor()
            
            sql = (
                "SELECT Flag, MkOrdNo "
                "FROM prdMkOrdMain "
                "WHERE MkOrdDate BETWEEN " + dateStart + " AND " + dateEnd + " AND Flag in (1, 2, 3)")
            manufactureOrders = cursorChi.execute(sql).fetchall()
            nCount = len(manufactureOrders)
            nDone = 0
            for manufactureOrder in manufactureOrders:
                self._migrate_inWorkorders(cursorChi, cursorBatom, manufactureOrder.MkOrdNo, manufactureOrder.Flag)
                nDone += 1
                if nDone % 10 == 0:
                    print str(nDone) + '/' + str(nCount)
                    self.env.cr.commit()
            self.env.cr.commit()
            connChi.close()
            connBatom.close()
        except Exception:
            _logger.warning('Exception in migrate_workorders_by_date:', exc_info=True)
            import pdb; pdb.set_trace()
            if connChi:
                connChi.close()
            if connBatom:
                connBatom.close()
    
    _ignored_sheets = [u'Sheet1', u'刀具資料-舊不用', u'酉潤刀數據', u'回明細', u'刀具清單']
    
    def _get_or_add_cutter_group(self, group_name):
        cutter_group = False
        cutter_groups = self.env['batom.cutter.group'].search([
            ('name', '=', group_name),
            ])
        if len(cutter_groups) > 0:
            cutter_group = cutter_groups[0]
        elif not group_name in self._ignored_sheets:
            try:
                cutter_group = self.env['batom.cutter.group'].create({
                    'name': group_name
                    })
            except Exception:
                _logger.warning('Exception in migrate_workorders_by_date:', exc_info=True)
                import pdb; pdb.set_trace()
        return cutter_group
        
    _sheet_format_column_mapping = ({
        u'圖面': 'image_file',
        u'狀態': 'state',
        u'本土編號': 'batom_code',
        u'履歷表': 'history_list',
        u'詢/訂價編號': 'inquiry_number',
        u'工件編號': 'product_code',
        u'刀具製造商': 'supplier',
        u'刀具種類': 'cutter_class',
        u'刀具編號': 'cutter_code',
        u'Material': 'material',
        u'TYPE': 'type',
        u'MOD': 'mod',
        u'DP': 'dp',
        u'PA': 'pa',
        u'Teeth': 'teeth',
        u'Teeth(工件)': 'teeth',
        u'OD': 'od',
        u'Length': 'length',
        u'Bore': 'bore',
        u'D+F': 'df',
        u'D+F(工件)': 'df',
        u'coating': 'coating',
        u'單價': 'price',
        u'匯率': 'exchange_rate',
        u'稅': 'tax',
        u'運費': 'shipping',
        u'修刀': 'dressing_cost',
        u'修刀費': 'dressing_cost',
        u'修刀費用': 'dressing_cost',
        u'磨刀費': 'sharpening_cost',
        u'鍍鈦費': 'titanium_cost',
        u'年份': 'year',
        u'Total': 'total',
        u'保管廠商': 'consigned_to',
        u'保管日期': 'consigned_date',
        u'歸還日期': 'returned_date',
        u'本土保管處': 'storage',
        u'本土保管處(2015/12/24)': 'storage',
        u'備註': 'remarks',
        u'詢價單': 'inquiry_form',
        u'供應商報價單': 'supplier_quotation',
        u'採購申請單': 'purchase_request',
        u'訂購單': 'purchase_order',
        u'訂單確認': 'order_confirmation',
        u'Invoice': 'invoice',
        u'訂購日期': 'order_date',
        u'期望交期': 'expected_delivery_date',
        u'現況': 'status',                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                    
        })
    def _getSheetFormatColumnName(self, sheet_title, column_title):
        column_name = False
        if column_title:
            if column_title in self._sheet_format_column_mapping:
                column_name = self._sheet_format_column_mapping[column_title]
            else:
                print u'column mapping not found: ' + sheet_title + u'/' + column_title
        return column_name
        
    _specialSheetNames = [u'攻牙刀&軸承', u'Liebherr資料']
    
    def _getSheetFormat(self, ws):
        special_format_name = False
        column_names = []
        original_column_names = []
        if ws.title in self._specialSheetNames:
            special_format_name = ws.title
        else:
            for cell in ws[1]:
                if cell.value:
                    column_names.append(self._getSheetFormatColumnName(ws.title, cell.value))
                    original_column_names.append(cell.value)
                else:
                    break
        return _tool_sheet_format(special_format_name, column_names, original_column_names)
        
    def _getProductIds(self, product_codes):
        product_list = re.findall(r"[\w]+", product_codes)
        product_ids = []
        for product_code in product_list:
            products = self.env['product.product'].search([
                ('default_code', '=', product_code),
                ])
            if len(products) == 0:
                suppliers = self.env['product.product'].search([
                    ('default_code', 'like', product_code),
                    ])
            if len(products) > 0:
                product_ids.append(products[0].id)
        return product_ids
    
    def _getSupplierId(self, supplier_name):
        suppliers = self.env['res.partner'].search([
            ('supplier', '=', True),
            ('is_company', '=', True),
            ('name', '=', supplier_name),
            ])
        if len(suppliers) == 0:
            suppliers = self.env['res.partner'].search([
                ('supplier', '=', True),
                ('is_company', '=', True),
                ('name', 'like', supplier_name),
                ])
        if len(suppliers) > 0:
            return suppliers[0].id
        else:
            return False
    
    _currency_format_mapping = ({
        u'[$\xa3-809]': 'GBP',
        u'[$\xa5-411]': 'JPY',
        u'"US$"': 'USD',
        u'"NT$"': 'TWD',
        })

    def _getCurrencyFromFormat(self, number_format):
        currency_id = _currencyIdConversion(self, 'TWD')
        for key, value in self._currency_format_mapping.iteritems():
            if number_format.find(key) >= 0:
                currency_id = _currencyIdConversion(self, value)
                break
        return currency_id
        
    def _appendRemarks(selv, values, original_column_name, value):
        if 'remarks' in values:
            values['remarks'] += '\n' + original_column_name + '=' + value
        else:
            values['remarks'] = original_column_name + '=' + value
            
    _hyperlink_columns = ([
        'image_file',
        'history_list',
        'inquiry_form',
        'supplier_quotation',
        'purchase_request',
        'purchase_order',
        'order_confirmation',
        ])
    _state_mapping = ({
        u'物料': 'material',
        u'進貨': 'bought',
        u'進貨-客供': 'consigned',
        u'進貨-轉賣': 'sold',
        })
    _number_columns = ([
        'exchange_rate',
        'total',
        ])
    _currency_columns = ([
        'price',
        'tax',
        'shipping',
        'dressing_cost',
        'sharpening_cost',
        'titanium_cost',
        ])
    
    def _addToolRow(self, cutter_group, sheet_format, row):
        cutter = False
        try:
            if not sheet_format.special_format_name:
                values = {'cutter_group_id': cutter_group.id}
                i = 0
                while i < len(sheet_format.column_names):
                    column_name = sheet_format.column_names[i]
                    original_column_name = sheet_format.original_column_names[i]
                    value = row[i].value
                    if column_name and value:
                        if column_name in self._hyperlink_columns and row[i].hyperlink:
                            value = row[i].hyperlink.target
                        elif column_name == 'state':
                            if value in self._state_mapping:
                                value = self._state_mapping[value]
                            else:
                                print u'invalid state: ' + cutter_group.name + u'/' + value
                                self._appendRemarks(values, original_column_name, (value if row[i].data_type == 's' else str(value)))
                                value = None
                        elif column_name == 'product_code':
                            product_ids = self._getProductIds(value if row[i].data_type == 's' else str(value))
                            if product_ids:
                                values['product_ids'] = [(4, product_ids)]
                        elif column_name == 'supplier':
                            supplier_id = self._getSupplierId(value if row[i].data_type == 's' else str(value))
                            if supplier_id:
                                values['supplier_ids'] = [(4, [supplier_id])]
                        elif column_name == 'consigned_to':
                            supplier_id = self._getSupplierId(value if row[i].data_type == 's' else str(value))
                            if supplier_id:
                                values['consigned_to_id'] = supplier_id
                        elif column_name in self._number_columns:
                            if not self._isQuantity(row[i]):
                                self._appendRemarks(values, original_column_name, (value if row[i].data_type == 's' else str(value)))
                                value = None
                        elif column_name in self._currency_columns:
                            if not self._isQuantity(row[i]):
                                idx =  value.find('RMB')
                                if idx >= 0:
                                    value = value[0:idx] + value[idx + 3:]
                                    try:
                                        value = float(s)
                                        currency_id = _currencyIdConversion(self, 'CNY')
                                        values[column_name + '_currency_id'] = currency_id;
                                    except ValueError:
                                        self._appendRemarks(values, original_column_name, (value if row[i].data_type == 's' else str(value)))
                                        value = None
                                else:
                                    self._appendRemarks(values, original_column_name, (value if row[i].data_type == 's' else str(value)))
                                    value = None
                            else:
                                currency_id = self._getCurrencyFromFormat(row[i].number_format)
                                if currency_id:
                                    values[column_name + '_currency_id'] = currency_id;
                        if value:
                            if column_name == 'remarks':
                                if 'remarks' in values:
                                    values['remarks'] = (value if row[i].data_type == 's' else str(value)) + '\n' + values['remarks']
                                else:
                                    values['remarks'] = value
                            else:
                                values[column_name] = value
                    elif original_column_name and value:
                        self._appendRemarks(values, original_column_name, (value if row[i].data_type == 's' else str(value)))
                    i += 1
                cutter = self.env['batom.cutter'].create(values)
        except Exception:
            _logger.warning('Exception in load_tool_sheet:', exc_info=True)
            import pdb; pdb.set_trace()
        return cutter
    
    def _addToolSheet(self, ws):
        print ws.title
        cutter_group = self._get_or_add_cutter_group(ws.title)
        if cutter_group:
            sheet_format = self._getSheetFormat(ws)
            first_row = True
            nDone = 0
            last_seconds = float(int(round(time.time() * 1000))) / 1000
            for row in ws.iter_rows():
                if first_row:
                    first_row = False
                else:
                    cutter = self._addToolRow(cutter_group, sheet_format, row)
                current_seconds = float(int(round(time.time() / 1000))) * 1000
                nDone += 1
                last_seconds = current_seconds
                if nDone % 10 == 0:
                    if nDone % 100 == 0:
                        print str(nDone) + " - " + str(current_seconds - last_seconds)
                    self.env.cr.commit()
            
    @api.multi
    def load_tool_sheet(self, xlsx_file):
        self.ensure_one()
        try:
            wb = load_workbook(filename = xlsx_file)
            for ws in wb.worksheets:
                self._addToolSheet(ws);
                self.env.cr.commit()
        except Exception:
            _logger.warning('Exception in load_tool_sheet:', exc_info=True)
            import pdb; pdb.set_trace()
    
class BatomPartnerMigration(models.Model):
    _name = 'batom.partner_migration'
    _description = 'Partner Data Migration'
    
    type = fields.Selection([(1, 'Customer'), (2, 'Sppplier')], string='Type', required=True)
    in_id = fields.Char('入料 Id', required=False, size=20)
    in_short_name = fields.Char('入料 Short Name', required=False, size=12)
    in_full_name = fields.Char('入料 Full Name', required=False, size=80)
    chi_id = fields.Char('正航 Id', required=False, size=20)
    chi_short_name = fields.Char('正航 Short Name', required=False, size=12)
    chi_full_name = fields.Char('正航 Full Name', required=False, size=80)
    odoo_id = fields.Char('Odoo Id', required=False, size=20)
    odoo_short_name = fields.Char('Odoo Short Name', required=False, size=12)
    odoo_full_name = fields.Char('Odoo Full Name', required=False, size=80)
    
class BatomMaterialAssignment(models.Model):
    _name = 'batom.material_assignment'
    _description = 'Raw Material Assignment'
    
    product_code = fields.Char('Product Code', size=20)
    product_qty = fields.Float('Product Quantity')
    production_code = fields.Char('Production Code', size=20)
    inventory_flag = fields.Char('Inventory In Flag', size=20) 
    inventory_code = fields.Char('Inventory In Code', size=20)
    adjustment_code = fields.Char('Inventory Adjustment Code', size=20)
    material_cost = fields.Float('Row Material Cost')
    cut_cost = fields.Float('Cut Cost')
    forge_cost = fields.Float('Forge Cost')
    material_lot = fields.Char('Material Lot', size=20)
    dispatch_date = fields.Date('Dispatch Date')
    imported = fields.Boolean('Imported', default=False)
    used = fields.Boolean('Used', default=False)
    used_by_production_ids = fields.Many2many(
        comodel_name='mrp.production', relation='batom_material_assignment_production_rel',
        column1='material_assignment_id', column2='production_id', string='Used by')
